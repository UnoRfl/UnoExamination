#!/usr/bin/env python3
"""Opens the workbook our JS writer produced with a real Excel parser.

A round-trip through our own reader only proves the two halves agree. This
proves the bytes are a workbook Excel, LibreOffice and Google Sheets will open.
Run after `node --test tests/xlsx.test.js`, which writes the fixture.
"""
import sys, datetime, pathlib
from openpyxl import load_workbook

f = pathlib.Path(__file__).parent / "out" / "sample.xlsx"
if not f.exists():
    sys.exit("missing fixture: run `node --test tests/xlsx.test.js` first")

fails = []
def check(label, ok, extra=""):
    print(("PASS " if ok else "FAIL ") + label + ("  " + str(extra) if extra and not ok else ""))
    if not ok: fails.append(label)

wb = load_workbook(f)
check("openpyxl opens the file", True)
check("both sheets are present", wb.sheetnames == ["Grades", "Per question"], wb.sheetnames)

ws = wb["Grades"]
check("dimensions are right", (ws.max_row, ws.max_column) == (5, 6), (ws.max_row, ws.max_column))
check("header text reads back", [c.value for c in ws[1]][:3] == ["Student", "E-mail", "Student no"])
check("the header is styled bold on the brand colour",
      ws["A1"].font.b is True and ws["A1"].fill.fgColor.rgb == "FF7A0D1F",
      (ws["A1"].font.b, ws["A1"].fill.fgColor.rgb))
check("a score is a number, not text", isinstance(ws["D2"].value, (int, float)), type(ws["D2"].value))
check("a percent keeps its format", "%" in (ws["E2"].number_format or ""), ws["E2"].number_format)
check("a date is a real datetime", isinstance(ws["F2"].value, datetime.datetime), ws["F2"].value)
check("that datetime is the one we wrote",
      isinstance(ws["F2"].value, datetime.datetime) and ws["F2"].value.date() == datetime.date(2026, 1, 15),
      ws["F2"].value)
check("a student who never submitted has a blank, not a zero", ws["F4"].value is None, repr(ws["F4"].value))
check("quotes survive the XML", ws["A3"].value == 'O\'Brien, "Mae"', ws["A3"].value)
check("angle brackets survive", ws["A4"].value == "Ng, Wei <lab>", ws["A4"].value)
check("non-ASCII survives", ws["A5"].value == "Ünlü, Zoë — ß", ws["A5"].value)
check("the header row is frozen", ws.freeze_panes == "A2", ws.freeze_panes)
check("the filter covers the data", ws.auto_filter.ref is not None, ws.auto_filter.ref)
check("column widths were applied", ws.column_dimensions["A"].width == 22, ws.column_dimensions["A"].width)

print()
print(f"=== {len(fails)} failed ===" if fails else "=== all checks passed ===")
sys.exit(1 if fails else 0)
